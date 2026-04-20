import openpyxl

wb = openpyxl.load_workbook('sheet.xlsx', data_only=False)

def print_headers_and_formulas(sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found")
        return
    ws = wb[sheet_name]
    print(f"--- {sheet_name} ---")
    
    headers = []
    # Assuming row 4 is the header row based on the PNL screenshot
    # Or maybe row 1 for Month vise
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5)):
        row_vals = [str(cell.value) if cell.value else '' for cell in row]
        if any(row_vals):
            print(f"Row {row_idx+1}: {row_vals[:20]}")
    
    print("\nSample Formulas from first data row:")
    for col_idx in range(1, 20):
        cell = ws.cell(row=5, column=col_idx) # Assuming row 5 is first data row for PNL
        if cell.value:
            print(f"Col {col_idx} ({cell.column_letter}): {cell.value}")

print_headers_and_formulas('PNL')
print_headers_and_formulas('Month vise')
